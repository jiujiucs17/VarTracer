import inspect
import os
import subprocess
import json
import xlsxwriter
from tqdm import tqdm
from datetime import datetime
import openpyxl
from openpyxl.comments import Comment
import re

def safe_serialize(obj):
        """将对象转换为字符串表示"""
        try:
            return str(obj)
        except Exception:
            return "<unserializable>"
        
def create_event(event_type, base_info, extra_info=None):
    """创建一个事件字典"""
    event = {"type": event_type, "details": base_info}
    if extra_info:
        event["details"].update(extra_info)
    return event

class FrameSnapshot:
    def __init__(self, frame):
        self.file_name = frame.f_code.co_filename
        self.function_name = frame.f_code.co_name
        self.line_no = frame.f_lineno
        self.locals = frame.f_locals.copy()
        self.globals = {k: str(v) for k, v in frame.f_globals.items() if k in frame.f_code.co_names}
        self.code_context = self._get_code_context(frame)

        self.package_name = frame.f_globals.get('__package__', None)
        self.module_name = frame.f_globals.get('__name__', None) 

    def _get_code_context(self, frame, context_lines=2):
        """Get the source code context around the current line."""
        try:
            lines, start = inspect.getsourcelines(frame.f_code)
            index = frame.f_lineno - start
            lower = max(index - context_lines, 0)
            upper = min(index + context_lines + 1, len(lines))
            return [line.rstrip('\n') for line in lines[lower:upper]]
        except (OSError, TypeError):
            return []

    def to_dict(self):
        """Convert the FrameSnapshot to a dictionary for easier serialization."""
        return {
            "filename": self.file_name,
            "function_name": self.function_name,
            "line_no": self.line_no,
            "locals": self.locals,
            "globals": self.globals,
            "code_context": self.code_context,
        }

    def __repr__(self):

        """Return a string representation of the FrameSnapshot."""
        return f"<FrameSnapshot {self.function_name} at {self.file_name}:{self.line_no}>"
    
def extension_interface(file_path, print=False):
    """
    不直接修改 file_path 指向的文件，而是在同目录下创建一个新文件（文件名加(VarTracerTemp)），
    写入插桩后的内容，运行新文件，生成数据，最后删除新文件。
    输出结果中所有的"(VarTracerTemp)"字符串都会被去除。
    """
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"The file {file_path} does not exist.")
    if not file_path.endswith('.py'):
        raise ValueError("The provided file is not a Python (.py) file.")

    file_dir = os.path.dirname(file_path)
    file_base, file_ext = os.path.splitext(os.path.basename(file_path))
    temp_file_name = f"{file_base}(VarTracerTemp){file_ext}"
    temp_file_path = os.path.join(file_dir, temp_file_name)
    result_path = os.path.join(file_dir, "result.json")

    def remove_vartracertemp_str(obj):
        """递归地将对象中的'(VarTracerTemp)'字符串去除"""
        if isinstance(obj, str):
            return obj.replace("(VarTracerTemp)", "")
        elif isinstance(obj, dict):
            return {remove_vartracertemp_str(k): remove_vartracertemp_str(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [remove_vartracertemp_str(i) for i in obj]
        else:
            return obj

    try:
        # 读取原始文件内容
        with open(file_path, 'r') as f:
            original_content = f.readlines()

        # === 重要说明 ===
        # 假设：被读取的原始文件中的所有 import 语句都在文件开头，且没有穿插主代码。
        # 这样做的原因是：sys.settrace 必须在所有 import 语句之后调用，否则会递归追踪 importlib 导致爆栈。
        # 因此，我们扫描原始文件的前若干行，将所有以 'import ' 或 'from ' 开头的行视为 import 语句，
        # 并把与 VarTracer 相关的插桩代码插入到这些 import 语句之后、主代码之前。
        # 这样可以最大程度保证追踪主代码和第三方库调用，但不会递归追踪 import 过程。

        # 找到 import 语句的结束行号
        import_end = 0
        for idx, line in enumerate(original_content):
            striped = line.lstrip()
            if striped.startswith('import ') or striped.startswith('from '):
                import_end = idx + 1
            elif striped == '' or striped.startswith('#'):
                continue  # 跳过空行和注释
            else:
                break  # 第一个非 import/空行/注释的地方停止

        # 构造插桩内容
        vartracer_instrument = [
            "from VarTracer import *\n",
            "import json\n",
            "vt = VarTracer()\n",
            "vt.start()\n"
        ]
        # 插入到所有 import 语句之后
        modified_content = (
            original_content[:import_end] +
            vartracer_instrument +
            original_content[import_end:] +
            [
                "\n",
                "vt.stop()\n",
                "exec_stack_json = vt.exec_stack_json()\n",
                "dep_tree = DependencyTree(call_stack=json.dumps(exec_stack_json))\n",
                "dep_dic = dep_tree.parse_dependency()\n",
                "\n",
                "result_json = {\n",
                "    'exec_stack': exec_stack_json,\n",
                "    'dependency': dep_dic\n",
                "}\n",
                "\n",
                f"with open(r'{result_path}', 'w') as result_file:\n",
                "    json.dump(result_json, result_file)\n"
            ]
        )

        # 写入临时文件
        with open(temp_file_path, 'w') as f:
            f.writelines(modified_content)

        # 运行临时文件
        try:
            std_exe_result = subprocess.run(['python3', temp_file_path],
                                            check=True,
                                            capture_output=True,
                                            text=True)
        except subprocess.CalledProcessError as e:
            raise e

        with open(result_path, 'r') as result_file:
            result_json = json.load(result_file)

        # 修正 dependency 中 temp_file_path 指向脚本的变量行号
        def fix_dependency_line_numbers(dependency, target_path):
            if target_path in dependency:
                for var, info in dependency[target_path].items():
                    if "lineNumber" in info and info["lineNumber"].isdigit():
                        info["lineNumber"] = str(int(info["lineNumber"]) - 4)
                    if "results" in info and isinstance(info["results"], dict):
                        for res in info["results"].values():
                            if "first_occurrence" in res and res["first_occurrence"].isdigit():
                                res["first_occurrence"] = str(int(res["first_occurrence"]) - 4)
                            if "co_occurrences" in res and isinstance(res["co_occurrences"], list):
                                res["co_occurrences"] = [
                                    str(int(x) - 4) if x.isdigit() else x for x in res["co_occurrences"]
                                ]

        # 修正 execution_stack 中 temp_file_path 指向脚本的行号，以及删除最后一个行的 VarTracer 相关信息，即代码内容为“exec_stack_json = vt.exec_stack_json()”的元素
        def fix_stack_line_numbers(execution_stack, target_path):
            for frame in execution_stack:
                details = frame.get("details", {})
                if details.get("file_path") == target_path:
                    if "line_no" in details and str(details["line_no"]).isdigit():
                        details["line_no"] = str(int(details["line_no"]) - 4)
                    if "line_content" in details and str(details["line_content"]).startswith("exec_stack_json = vt.exec_stack_json()"):
                        execution_stack.remove(frame)

            

        target_path = temp_file_path
        if "dependency" in result_json:
            fix_dependency_line_numbers(result_json["dependency"], target_path)

        if "execution_stack" in result_json.get("exec_stack", {}):
            fix_stack_line_numbers(result_json["exec_stack"]["execution_stack"], target_path)

        # 清理 tracing 代码带来的额外项
        exec_stack = result_json.get("exec_stack", {})
        if "execution_stack" in exec_stack:
            filtered_stack = []
            for frame in exec_stack["execution_stack"]:
                details = frame.get("details", frame)
                module = details.get("module") or details.get("module_name")
                if not (module and str(module).startswith("VarTracer")):
                    filtered_stack.append(frame)
            exec_stack["execution_stack"] = filtered_stack

        dependency = result_json.get("dependency", {})
        for dep_file in list(dependency.keys()):
            var_items = dependency[dep_file]
            keys_to_del = [k for k, v in var_items.items() if v.get("variableName") == "exec_stack_json"]
            for k in keys_to_del:
                del var_items[k]
            if dep_file.endswith("VarTracer_Code.py"):
                del dependency[dep_file]
        keys_to_del = [k for k in dependency.keys() if k.endswith("VarTracer_Core.py")]
        for k in keys_to_del:
            del dependency[k]

        # 去除所有(VarTracerTemp)字符串
        result_json = remove_vartracertemp_str(result_json)

        if print:
            print(json.dumps(result_json, indent=4))
        return result_json

    finally:
        # 删除临时文件和结果文件
        # if os.path.exists(result_path):
        #     os.remove(result_path)
        # if os.path.exists(temp_file_path):
        #     os.remove(temp_file_path)
        pass

def break_down_granularity(exec_stack, dep_tree):
    """
    将 exec_stack（JSON对象）展开为事件列表，并为每个事件添加 module_event 和 function_event 字段。
    每个事件包含如下字段：
      event_no, file_path, event_type, module, call_depth, line_no, line_content, variable, dep. variable,
      module_event, function_event
    """
    def get_dep_variables(dep_tree, event):
        assigned_vars = event.get("details", {}).get("assigned_vars", [])
        dep_var_inline = event.get("details", {}).get("dependencies", [])
        event_file_path = event.get("details", {}).get("file_path", "")

        event_module = os.path.splitext(os.path.basename(event_file_path))[0]
        event_func = event.get("details", {}).get("func", "")
        text = ""
        if not assigned_vars or len(assigned_vars) == 0:
            return ""
        else:
            for var in assigned_vars:
                # 作用域链格式：filename.toplevelscopename.**.parentscopename.varname
                scoped_var = f"{event_module}.{event_func}.{var}"
                dependent_vars_json = dep_tree.get(event_file_path, {}).get(scoped_var, {}).get("results", {})
                dependent_vars = []
                for dep_var, dep_info in dependent_vars_json.items():
                        dependent_vars.append(dep_var)

                text_dep_vars = f"{var}: INLINE DEPENDENCIES: "
                text_dep_vars += f"{', '.join(dep_var_inline)}" if len(dep_var_inline)>0 else "NONE"
                text_dep_vars += f" ｜-|-｜ ALL DEPENDENCIES (within file): {', '.join(dependent_vars)}" if len(dependent_vars)>0 else "NONE"
                text += text_dep_vars + " ｜-|-|-｜ "
            
            return text.strip(" ｜-|-|-｜ ")

    # 展平 execution_stack
    events = []
    def traverse(stack, event_no_counter):
        for event in stack:
            details = event.get("details", {})
            # 变量名
            assigned_vars = details.get("assigned_vars", [])
            variable = assigned_vars[0] if assigned_vars else ""
            # 依赖变量
            dep_variable = get_dep_variables(dep_tree, event)

            # if module name is "(unknown-module)", use string "top level script" instead
            if details.get("module") == "(unknown-module)":
                module_name = "**** top level script ****"
            else:
                module_name = details.get("module")
            events.append({
                "event_no": event_no_counter[0],
                "file_path": details.get("file_path", ""),
                "event_type": event.get("type", ""),
                "module": module_name,
                "function": details.get("func", ""),
                "call_depth": details.get("depth", ""),
                "line_no": details.get("line_no", ""),
                "line_content": details.get("line_content", ""),
                "variable": variable,
                "dep.variable": dep_variable,
                # module_event/function_event 暂时空
            })
            event_no_counter[0] += 1
            # 递归 daughter_stack
            if "daughter_stack" in details:
                traverse(details["daughter_stack"], event_no_counter)
    traverse(exec_stack.get("execution_stack", []), [1])

    # 计算 module_event 和 function_event
    module_count = {}
    function_count = {}
    last_module = None
    last_function = None
    module_event_idx = 0
    function_event_idx = 0

    for idx, event in enumerate(events):
        module = event["module"]
        function = event.get("function")

        # module_event
        if module != last_module:
            module_event_idx = module_count.get(module, 0) + 1
            module_count[module] = module_event_idx
            last_module = module
        event["module_event"] = f"{module}_{module_event_idx}"

        # function_event
        func_key = (f"{module}_{module_event_idx}", function)
        if func_key != last_function:
            function_event_idx = function_count.get(func_key, 0) + 1
            function_count[func_key] = function_event_idx
            last_function = func_key
        event["function_event"] = f"{module}_{module_event_idx}_{function}_{function_event_idx}"

    return events

def compare_event_lists(exec_stack0, exec_stack1, dep_tree0, dep_tree1):
    """
    对比两个 exec_stack，分别展开为事件列表，并为每个事件添加 unique_to_this_feature 字段。
    如果某事件仅在 events1 或 events2 中出现，则其 unique_to_this_feature 字段为 True。
    判断标准为 (file_path, line_no, line_content) 三元组完全一致。
    返回 (events1, events2)
    """
    events0 = break_down_granularity(exec_stack0, dep_tree0)
    events1 = break_down_granularity(exec_stack1, dep_tree1)

    def make_event_key(event):
        return (event.get("file_path", ""), str(event.get("line_no", "")), str(event.get("line_content", "")))

    set1 = set(make_event_key(e) for e in events1)
    set0 = set(make_event_key(e) for e in events0)

    for e in events0:
        e["unique_to_this_feature"] = make_event_key(e) not in set1
    for e in events1:
        e["unique_to_this_feature"] = make_event_key(e) not in set0

    return events0, events1

def compare_exec_stack(exec_stack0, exec_stack1, dep_tree0, dep_tree1, output_path):
    """
    对比两个 exec_stack，分别以 module granularity、function granularity、event granularity 展示。
    输出为 xlsx 文件，包含多个 sheet：
      1. module_granularity：每个 module_event 的信息
      2. module_event_n：每个 module_event 的 func_event 信息，包含所有 func_event，并有 this_context 字段
      3. module_event_n_func_event_m：每个 func_event 的 line_event 信息，包含所有 line_event，并有 this_context 字段
    所有 sheet 都包含完整数据，并添加 this_context 字段。
    保存后用 openpyxl 设置除 module_granularity 外所有 sheet 的筛选条件为 this_context=true。
    """

    # 1. 生成 events0, events1
    events0, events1 = compare_event_lists(exec_stack0, exec_stack1, dep_tree0, dep_tree1)

    def build_json(events):
        # 提取所有 func_event
        func_event_map = {}
        func_event_list = []
        for e in events:
            fe = e["function_event"]
            if fe not in func_event_map:
                func_event_map[fe] = []
            func_event_map[fe].append(e)

        for fe, fe_events in func_event_map.items():
            func_name = fe_events[0].get("function_name") or fe_events[0].get("function") or ""
            file_path = fe_events[0].get("file_path", "")
            fe_event_nos = [ev["event_no"] for ev in fe_events]
            fe_unique_to_this_feature = any(ev["unique_to_this_feature"] for ev in fe_events)
            line_events = []
            for lev in fe_events:
                line_events.append({
                    "event_no": lev["event_no"],
                    "event_type": lev["event_type"],
                    "call_depth": lev["call_depth"],
                    "file_path": lev["file_path"],
                    "line_no": lev["line_no"],
                    "line_content": lev["line_content"],
                    "variable": lev["variable"],
                    "dep.variable": lev["dep.variable"],
                    "unique_to_this_feature": lev["unique_to_this_feature"],
                    "module_event": lev["module_event"],
                    "function_event": fe
                })
            func_event_list.append({
                "func_event_no": len(func_event_list) + 1,
                "func_event_identifier": fe, #fe_events[0]["function_event"],
                "func_name": func_name,
                "file_path": file_path,
                "first_event_no": min(fe_event_nos),
                "last_event_no": max(fe_event_nos),
                "unique_to_this_feature": fe_unique_to_this_feature,
                "line_events": line_events,
                "module_event": fe_events[0]["module_event"]
            })
        # 提取所有 module_event
        module_event_map = {}
        module_event_list = []
        for e in events:
            me = e["module_event"]
            if me not in module_event_map:
                module_event_map[me] = []
            module_event_map[me].append(e)
        for idx, (me, me_events) in enumerate(module_event_map.items(), 1):
            module_name = me_events[0]["module"]
            event_nos = [ev["event_no"] for ev in me_events]
            unique_to_this_feature = any(ev["unique_to_this_feature"] for ev in me_events)
            module_event_list.append({
                "module_event_no": idx,
                "module_event_identifier": me,
                "module_name": module_name,
                "first_event_no": min(event_nos),
                "last_event_no": max(event_nos),
                "unique_to_this_feature": unique_to_this_feature
            })
        return module_event_list, func_event_list

    def write_xlsx(events, filename):
        module_event_list, func_event_list = build_json(events)
        # 提取所有 line_event
        all_line_events = []
        for func in func_event_list:
            for le in func["line_events"]:
                all_line_events.append(le)
        workbook = xlsxwriter.Workbook(os.path.join(output_path, filename))
        highlight_cell_format_pale = workbook.add_format({
            'bg_color': "#FFEDEF",    # 背景色
            'font_color': '#9C0006',  # 字体颜色
            'border': 1,  # 边框
            'align': 'left', 
            'valign': 'vcenter',
            'align': 'center'
        })
        highlight_cell_format_shine = workbook.add_format({
            'bg_color': "#FFC7CE",    # 背景色
            'font_color': '#9C0006',  # 字体颜色
            'border': 1,  # 边框
            'align': 'left', 
            'valign': 'vcenter',
            'align': 'center'

        })
        regular_cell_format = workbook.add_format({
            'bg_color': '#FFFFFF',    # 背景色
            'font_color': '#000000',  # 字体颜色
            'border': 1,  # 边框
            'align': 'left', 
            'valign': 'vcenter',
            'align': 'center'
        })
        hyperlink_format = workbook.add_format({
            'font_color': '#0000FF',  # 超链接字体颜色
            'underline': 1,           # 下划线
            'align': 'left', 
            'valign': 'vcenter',
            'align': 'left'
        })
        # 1. module_granularity sheet
        module_headers = [
            "module_event_no", "module_name", "execution_sequence_slice", "unique_to_this_feature", "associated_func_events" 
        ]
        module_sheet = workbook.add_worksheet("module_granularity")
        for col, h in enumerate(module_headers):
            module_sheet.write(0, col, h)

        for row, module in enumerate(module_event_list, 1):
            mrow_format = regular_cell_format
            if module["unique_to_this_feature"]:
                mrow_format = highlight_cell_format_shine

            for col, key in enumerate(module_headers[:-1]):  # 最后一个字段 "associated_func_events" 不在 module_event_list 中
                if key == "execution_sequence_slice":
                    module_sheet.write(row, col, f"No. {module['first_event_no']} to {module['last_event_no']} of all exec events", mrow_format)
                else:
                    module_sheet.write(row, col, module[key], mrow_format)

        # 2. 每个 module_event_n sheet
        for row, module in tqdm(enumerate(module_event_list, 1), desc="Generating module_events & func_events", total=len(module_event_list)):
            func_headers = [
            "func_event_no", "func_name", "file_path", "module_event_no", "execution_sequence_slice", "unique_to_this_feature", f"within_module_event_{row}", "associated_line_events"
        ]

            sheet_name = f"module_event_{module['module_event_no']}"
            func_sheet = workbook.add_worksheet(sheet_name)

            if len(func_event_list) == 0:
                # 如果没有 func_event，则跳过这个 module_event 的 sheet
                module_sheet.write(row, len(module_headers) - 1, "-")
                continue

            module_sheet.write_url(row, len(module_headers) - 1, 
                                   f"internal:'{sheet_name}'!A1",
                                   string=f"func_events")  # 添加超链接到 func_event sheet

            for col, h in enumerate(func_headers):
                func_sheet.write(0, col, h)
            
            for frow, func in enumerate(func_event_list, 1):
                this_context = func["module_event"] == module["module_event_identifier"]
                if this_context and func["unique_to_this_feature"]:
                    frow_format = highlight_cell_format_shine
                elif not this_context and func["unique_to_this_feature"]:
                    frow_format = highlight_cell_format_pale
                else:
                    frow_format = regular_cell_format

                for col, key in enumerate(func_headers[:-1]):  # 最后一个字段 "associated_line_events" 不在 func_event_list 中
                    if key == f"within_module_event_{row}":
                        func_sheet.write(frow, col, this_context, frow_format)
                    elif key == "execution_sequence_slice":
                        func_sheet.write(frow, col, f"No. {func['first_event_no']} to {func['last_event_no']} of all exec events", frow_format)
                    elif key == "module_event_no":
                        # 查询 identifier为 module_event 的 module_event_no
                        module_event_no = next((me["module_event_no"] for me in module_event_list if me["module_event_identifier"] == func["module_event"]), None)
                        func_sheet.write(frow, col, module_event_no, frow_format)
                    else:
                        func_sheet.write(frow, col, func[key] if key in func else "", frow_format)

                # 3. 每个 module_event_n_func_event_m sheet
                if this_context:
                    event_headers = [
                        "event_no", "event_type", "call_depth", "file_path", "module_event_no", "function_event_no", "line_no",
                        "line_content", "variable", "dep.variable", "unique_to_this_feature", f"within_func_event_{frow}"
                    ]
                    event_sheet_name = f"module_event_{module['module_event_no']}_func_event_{frow}"
                    event_sheet = workbook.add_worksheet(event_sheet_name)

                    func_sheet.write_url(frow, len(func_headers) - 1,
                                         f"internal:'{event_sheet_name}'!A1",
                                         string=f"line_events")

                    for ecol, eh in enumerate(event_headers):
                        event_sheet.write(0, ecol, eh)
                    for erow, event in enumerate(all_line_events, 1):
                        this_context_event = event["function_event"] == func["func_event_identifier"]
                        # erow_format = highlight_cell_format if this_context_event else regular_cell_format

                        if this_context_event and event["unique_to_this_feature"]:
                            erow_format = highlight_cell_format_shine
                        elif not this_context_event and event["unique_to_this_feature"]:
                            erow_format = highlight_cell_format_pale
                        else:
                            erow_format = regular_cell_format

                        for ecol, eh in enumerate(event_headers):
                            if eh == f"within_func_event_{frow}":
                                event_sheet.write(erow, ecol, this_context_event, erow_format)
                            elif eh == "module_event_no":
                                # 查询 identifier为 module_event 的 module_event_no
                                module_event_no = next((me["module_event_no"] for me in module_event_list if me["module_event_identifier"] == event["module_event"]), None)
                                event_sheet.write(erow, ecol, module_event_no, erow_format)
                            elif eh == "function_event_no":
                                # 查询 identifier为 func_event 的 func_event_no
                                func_event_no = next((fe["func_event_no"] for fe in func_event_list if fe["func_event_identifier"] == event["function_event"]), None)
                                event_sheet.write(erow, ecol, func_event_no, erow_format)
                            else:
                                value = "-"
                                if eh in event:
                                    value = event[eh]
                                    if value is None or value == "":
                                        value = "-"
                                event_sheet.write(erow, ecol, value, erow_format)

                    last_col = event_sheet.dim_colmax if hasattr(event_sheet, 'dim_colmax') and event_sheet.dim_colmax is not None else 0
                    last_row = event_sheet.dim_rowmax if hasattr(event_sheet, 'dim_rowmax') and event_sheet.dim_rowmax is not None else 0
                    # 在最后一列右侧第一个单元格添加超链接
                    event_sheet.write_url(0, last_col + 2, f"internal:'{sheet_name}'!A1", cell_format=hyperlink_format, string=f"back to function granularity")
                    event_sheet.write_url(last_row + 2, 0, f"internal:'{sheet_name}'!A1", cell_format=hyperlink_format, string=f"back to function granularity")
                else:
                    func_sheet.write(frow, len(func_headers) - 1, "-")
        
        print("cleaning up workbooks...")
        for worksheet in workbook.worksheets():
            # 设置冻结窗格
            worksheet.freeze_panes(1, 0)
            # 设置默认列宽
            default_width = 25
            worksheet.set_column('A:XFD', default_width)
            # 获取当前工作表的数据行数和列数
            last_row = worksheet.dim_rowmax if hasattr(worksheet, 'dim_rowmax') and worksheet.dim_rowmax is not None else 1
            last_col = worksheet.dim_colmax if hasattr(worksheet, 'dim_colmax') and worksheet.dim_colmax is not None else 0
            # 在最后一列右侧第一个单元格添加超链接
            worksheet.write_url(0, (last_col - 1 if len(worksheet.name) >= 22 else last_col + 1), "internal:'module_granularity'!A1", cell_format=hyperlink_format, string="back to module granularity")
            worksheet.write_url((last_row + 1 if len(worksheet.name) >= 22 else last_row + 2), 0, "internal:'module_granularity'!A1", cell_format=hyperlink_format, string="back to module granularity")

        workbook.close()

        # 用 openpyxl 设置筛选条件
        wb = openpyxl.load_workbook(os.path.join(output_path, filename))
        for ws in wb.worksheets:
            if ws.title == "module_granularity":
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == "unique_to_this_feature":
                        unique_to_this_feature_col = col
                        break

                ws.auto_filter.ref = ws.dimensions
                ws.auto_filter.add_filter_column(unique_to_this_feature_col - 1, ["True"])
            else:
                # 找到 this_context 列
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value.startswith("within_"):
                        this_context_col = col
                        break
                # 设置筛选
                ws.auto_filter.ref = ws.dimensions
                ws.auto_filter.add_filter_column(this_context_col - 1, ["True"])
            # 从左向右遍历第一行的单元格，遇到最后一个有内容的单元格时停止。使用每个单元格的内容作为键，来查询它们在label_meanings中的含义，并作为这个单元格的comment, 对于在label_meanings中没有的键，则comment为空字符串。
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                cell_value = re.sub(r"\d+", "", cell_value)
                cell_comment_string = ""
                if cell_value in label_meanings:
                    cell_comment_string = label_meanings[cell_value]
                cell_comment = Comment(cell_comment_string, "Label Meaning")
                cell_comment.width = 200
                cell_comment.height = 100
                ws.cell(row=1, column=col).comment = cell_comment

        wb.save(os.path.join(output_path, filename))

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    write_xlsx(events0, f"stack0_{timestamp}.xlsx")
    write_xlsx(events1, f"stack1_{timestamp}.xlsx")

    print(f"length of events0: {len(events0)}")
    print(f"length of events1: {len(events1)}")

    return None, None

def compare_dependency(dep1, dep2, output_path, output_name="compare_dependency_result.xlsx"):
    """
    比较两个 VarTracer.dep_tree_json 生成的依赖树（JSON 格式），
    输出两部分差异：
      1. 只在 A 中有但 B 中没有的文件/变量
      2. 只在 B 中有但 A 中没有的文件/变量
    并将结果保存为 xlsx 文件，包含两个 sheet，按文件路径、变量名、变量行号纵向合并单元格。
    """
    import xlsxwriter
    import os

    output_xlsx = os.path.join(output_path, output_name)

    def extract_vars(dep):
        """
        提取所有 (file_path, var_name, line_no, results) 元组。
        返回: {(file_path, var_name, line_no): results}
        """
        result = {}
        for file_path, var_dict in dep.items():
            if file_path == "trace_started_at":
                continue
            for var_key, var_info in var_dict.items():
                var_name = var_info.get("variableName", var_key)
                line_no = str(var_info.get("lineNumber", "-"))
                results = var_info.get("results", {})
                result[(file_path, var_name, line_no)] = results
        return result

    vars_A = extract_vars(dep1)
    vars_B = extract_vars(dep2)

    # 1. 只在 A 中有但 B 中没有
    only_in_A = {k: v for k, v in vars_A.items() if k not in vars_B}
    # 2. 只在 B 中有但 A 中没有
    only_in_B = {k: v for k, v in vars_B.items() if k not in vars_A}

    def prepare_sheet_data(var_dict):
        """
        生成 sheet 数据，每一行是
        (file_path, var_name, line_no, dep_var, first_occurrence, co_occurrences)
        """
        rows = []
        for (file_path, var_name, line_no), results in var_dict.items():
            if not results:
                # 没有依赖变量
                rows.append((file_path, var_name, line_no, "-", "-", "-"))
            else:
                for dep_var, dep_info in results.items():
                    first_occ = str(dep_info.get("first_occurrence", "-"))
                    co_occs = dep_info.get("co_occurrences", [])
                    if isinstance(co_occs, list):
                        co_occs_str = ", ".join(str(x) for x in co_occs) if co_occs else "-"
                    else:
                        co_occs_str = "-"
                    rows.append((file_path, var_name, line_no, dep_var, first_occ, co_occs_str))
        return rows

    data_A = prepare_sheet_data(only_in_A)
    data_B = prepare_sheet_data(only_in_B)

    # 写入 xlsx
    workbook = xlsxwriter.Workbook(output_xlsx)
    border_fmt = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
    header_fmt = workbook.add_format({'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter'})

    headers = ["File Path", "Variable", "Line No", "Dep.Var", "First Occurrence", "Co-occurrences"]

    def write_sheet(sheet_name, data):
        worksheet = workbook.add_worksheet(sheet_name)
        for col, h in enumerate(headers):
            worksheet.write(0, col, h, header_fmt)
        if not data:
            return

        from collections import defaultdict
        # 先按文件分组，再按变量名+行号分组
        file_group = defaultdict(list)
        for row in data:
            file_group[row[0]].append(row)

        row_idx = 1
        for file_path in sorted(file_group.keys()):
            file_rows = file_group[file_path]
            # 按变量名+行号分组
            var_group = defaultdict(list)
            for row in file_rows:
                var_group[(row[1], row[2])].append(row)
            var_items = sorted(var_group.items(), key=lambda x: (x[0][0], int(x[0][1]) if str(x[0][1]).isdigit() else 0))
            file_row_start = row_idx
            for (var_name, line_no), rows in var_items:
                n = len(rows)
                # 合并变量名和行号
                worksheet.merge_range(row_idx, 1, row_idx + n - 1, 1, var_name, border_fmt)
                worksheet.merge_range(row_idx, 2, row_idx + n - 1, 2, line_no, border_fmt)
                for i, row in enumerate(rows):
                    worksheet.write(row_idx + i, 3, row[3], border_fmt)
                    worksheet.write(row_idx + i, 4, row[4], border_fmt)
                    worksheet.write(row_idx + i, 5, row[5], border_fmt)
                row_idx += n
            # 合并文件名
            worksheet.merge_range(file_row_start, 0, row_idx - 1, 0, file_path, border_fmt)

    write_sheet("Only_in_A", data_A)
    write_sheet("Only_in_B", data_B)
    workbook.close()
    return output_xlsx


label_meanings = {
    "module_event_no": "Module Event number in the entire Execution Stack, representing the order of Module Events being executed.",
    "module_name": "Name of the module where the event occurred.",
    "execution_sequence_slice": "Slice of the execution sequence, indicating the range of Line Event numbers contained by this Event.",
    "unique_to_this_feature": "Indicates whether this Event only showed up in the script with this feature enabled.",
    "associated_func_events": "Hyperlink to the Function Event sheet, which shows all Function Events associated with this Module Event.",

    "func_event_no": "Function Event number in the entire Execution Stack, representing the order of Function Events being executed.",
    "func_name": "Name of the Function where the event occurred.",
    "file_path": "File path where the event occurred.",
    "within_module_event_": "Indicates whether this Function Event is within the context of the Module Event identified by the label.",
    "associated_line_events": "Hyperlink to the Line Event sheet, which shows all Line Events associated with this Function Event.",

    "event_no": "Event number in the entire Execution Stack, representing the order of Line Events being executed.",
    "event_type": "Type of the event, one of the following: 'LINE', 'CALL', 'RETURN', 'EXCEPTION'.",
    "call_depth": "Call depth of the event, indicating how deep the call stack is at this point.",
    "module_event_no": "Module Event number associated with this Line Event, conceptually linking it to the Module Event sheet.",
    "function_event_no": "Function Event number associated with this Line Event, conceptually linking it to the Function Event sheet.",
    "line_no": "Line number in the python file where the Line Event occurred.",
    "line_content": "Code content of the Line Event.",
    "variable": "Variable being modified or assigned value to with this Line Event, if any.",
    # considering changing the dep.variable meaning to "variables whose value **this line** depends on, blabla"
    "dep.variable": "Variables whose value the variable in the left depends on, showing inline dependencies and all dependencies within the file.",
    "within_func_event_": "Indicates whether this Line Event is within the context of the Function Event identified by the label.",

    "back to module granularity": "Hyperlink to the Module Granularity sheet, which shows all Module Events.",
    "back to function granularity": "Hyperlink to the Function Granularity sheet, which shows all Function Events. While highlighting the current context.",
}